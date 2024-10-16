import warnings
warnings.filterwarnings('ignore')

import os
import time
import easyocr
from PIL import Image, ImageFilter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm  # Import tqdm for progress bar
from spellchecker import SpellChecker
import string
import re

# Preprocessing function to clean and prepare words for spell checking
def preprocess_words(text):
    # Remove punctuation using regex
    text = re.sub(r'[^\w\s]', '', text)
    # Convert to lowercase
    words = text.lower().split()
    return words

def extract_english_text(text):
    # Remove Mandarin characters, keep English words
    return re.sub(r'[^\x00-\x7F]+', '', text).strip()

def extract_mandarin_text(text):
    # Extract Mandarin characters
    mandarin_characters = re.findall(r'[\u4e00-\u9fff]+', text)
    return ' '.join(mandarin_characters)

# Function to process images and extract text with verification
def extract_text_from_images(directory):
    # Initialize the OCR readers
    reader_simplified = easyocr.Reader(['en', 'ch_sim'], verbose=False)
    reader_traditional = easyocr.Reader(['en', 'ch_tra'], verbose=False)

    # Initialize spell checker
    spell = SpellChecker(language='en')
    custom_words = ['cas', 'CAS', 'CAS*']
    spell.word_frequency.load_words(custom_words)

    # Initialize list to store data for the excel file
    data = []
    total_images = 0
    images_with_text = 0
    images_without_text = 0
    failed_extractions = 0
    low_confidence_count = 0
    spell_errors_count = 0
    total_extracted_elements = 0  # Counter for total extracted elements

    # For sharpened images
    sharpened_low_confidence_count = 0
    sharpened_spell_errors_count = 0
    total_sharpened_extracted_elements = 0

    # To calculate average confidences
    total_confidence = 0
    total_sharpened_confidence = 0

    # Collect texts for text files
    english_texts = []
    mandarin_texts = []
    original_texts = []

    # Supported image extensions
    supported_extensions = ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')

    # Loop through the directory with tqdm progress bar
    for filename in tqdm(os.listdir(directory), desc="Processing Images", unit="image"):
        if filename.lower().endswith(supported_extensions):
            total_images += 1  # Count total images
            file_path = os.path.join(directory, filename)

            # Initialize data dictionary for this image
            image_data = {
                'file_path': file_path,
                'filename': filename,
            }

            # Open and read the image using EasyOCR
            try:
                # Process original image
                # Try using the Simplified Chinese reader first
                results = reader_simplified.readtext(file_path, detail=1)

                if not results:  # If no text found, try Traditional Chinese reader
                    results = reader_traditional.readtext(file_path, detail=1)

                if results:
                    images_with_text += 1  # Text was found
                    total_extracted_elements += len(results)  # Count extracted elements
                    extracted_text = ' '.join([res[1] for res in results])

                    # Calculate average confidence
                    confidences = [res[2] for res in results]
                    avg_confidence = sum(confidences) / len(confidences)
                    total_confidence += avg_confidence

                    # Count low confidence text elements
                    low_confidence_elements = [c for c in confidences if c < 0.5]
                    low_confidence_count += len(low_confidence_elements)

                    # Determine verification status based on average confidence
                    verification_status = "Low Confidence" if avg_confidence < 0.6 else "Verified"

                    # Spell Checking for English Text
                    misspelled_words = []
                    detected_language = 'en'  # Default language since we are focusing on English
                    if detected_language == 'en':
                        words = preprocess_words(extracted_text)
                        misspelled = spell.unknown(words)
                        if misspelled:
                            misspelled_words = list(misspelled)
                            spell_errors_count += len(misspelled)
                            verification_status += f" | {len(misspelled)} Spell Errors"

                    # Collect English and Mandarin texts
                    english_text = extract_english_text(extracted_text)
                    if english_text:
                        original_texts.append(extracted_text)
                    mandarin_text = extract_mandarin_text(extracted_text)
                    if mandarin_text:
                        mandarin_texts.append(mandarin_text)

                else:
                    images_without_text += 1  # No text found
                    extracted_text = "No text found"
                    avg_confidence = 0
                    verification_status = "No Text"
                    misspelled_words = []

                # Add data to image_data dictionary
                image_data['extracted_text'] = extracted_text
                image_data['avg_confidence'] = avg_confidence
                image_data['verification_status'] = verification_status
                image_data['misspelled_words'] = ', '.join(misspelled_words)

                # Now process the sharpened image
                # Open image
                img = Image.open(file_path)
                # Sharpen image
                sharpened_img = img.filter(ImageFilter.SHARPEN)
                # Save to a temporary file
                temp_sharpened_file = os.path.join(directory, 'temp_sharpened_' + filename)
                sharpened_img.save(temp_sharpened_file)

                # Process the sharpened image
                results_sharp = reader_simplified.readtext(temp_sharpened_file, detail=1)
                if not results_sharp:  # If no text found, try Traditional Chinese reader
                    results_sharp = reader_traditional.readtext(temp_sharpened_file, detail=1)

                if results_sharp:
                    total_sharpened_extracted_elements += len(results_sharp)
                    sharpened_extracted_text = ' '.join([res[1] for res in results_sharp])

                    # Calculate average confidence
                    sharpened_confidences = [res[2] for res in results_sharp]
                    sharpened_avg_confidence = sum(sharpened_confidences) / len(sharpened_confidences)
                    total_sharpened_confidence += sharpened_avg_confidence

                    # Count low confidence text elements
                    sharpened_low_confidence_elements = [c for c in sharpened_confidences if c < 0.5]
                    sharpened_low_confidence_count += len(sharpened_low_confidence_elements)

                    # Determine verification status based on average confidence
                    sharpened_verification_status = "Low Confidence" if sharpened_avg_confidence < 0.6 else "Verified"

                    # Spell Checking for English Text
                    sharpened_misspelled_words = []
                    detected_language = 'en'  # Default language since we are focusing on English
                    if detected_language == 'en':
                        words_sharp = preprocess_words(sharpened_extracted_text)
                        misspelled_sharp = spell.unknown(words_sharp)
                        if misspelled_sharp:
                            sharpened_misspelled_words = list(misspelled_sharp)
                            sharpened_spell_errors_count += len(misspelled_sharp)
                            sharpened_verification_status += f" | {len(misspelled_sharp)} Spell Errors"

                    # Collect English and Mandarin texts from sharpened images
                    english_text_sharp = extract_english_text(sharpened_extracted_text)
                    if english_text_sharp:
                        english_texts.append(english_text_sharp)
                    mandarin_text_sharp = extract_mandarin_text(sharpened_extracted_text)
                    if mandarin_text_sharp:
                        mandarin_texts.append(mandarin_text_sharp)

                else:
                    sharpened_extracted_text = "No text found"
                    sharpened_avg_confidence = 0
                    sharpened_verification_status = "No Text"
                    sharpened_misspelled_words = []
                    words_sharp = []

                # Remove the temporary sharpened image file
                os.remove(temp_sharpened_file)

                # Add data to image_data dictionary
                image_data['sharpened_extracted_text'] = sharpened_extracted_text
                image_data['sharpened_avg_confidence'] = sharpened_avg_confidence
                image_data['sharpened_verification_status'] = sharpened_verification_status
                image_data['sharpened_misspelled_words'] = ', '.join(sharpened_misspelled_words)

                # Get the sharpened extracted text without misspelled words
                words_sharp_corrected = [word for word in words_sharp if word not in sharpened_misspelled_words]
                sharpened_text_without_misspelled = ' '.join(words_sharp_corrected)
                image_data['sharpened_text_without_misspelled'] = sharpened_text_without_misspelled

                # Compare extracted_text and sharpened_extracted_text to find added and removed words
                words_original = set(preprocess_words(extracted_text))
                words_sharpened = set(preprocess_words(sharpened_extracted_text))
                added_words = words_sharpened - words_original
                removed_words = words_original - words_sharpened

                image_data['added_words'] = ', '.join(added_words)
                image_data['removed_words'] = ', '.join(removed_words)

                # Append image_data to data list
                data.append(image_data)

            except Exception as e:
                failed_extractions += 1  # OCR failed
                print(f"Error processing {filename}: {e}")
                image_data['extracted_text'] = "Error in OCR"
                image_data['avg_confidence'] = 0
                image_data['verification_status'] = "Failed"
                image_data['misspelled_words'] = ''
                image_data['sharpened_extracted_text'] = "Error in OCR"
                image_data['sharpened_avg_confidence'] = 0
                image_data['sharpened_verification_status'] = "Failed"
                image_data['sharpened_misspelled_words'] = ''
                image_data['sharpened_text_without_misspelled'] = ''
                image_data['added_words'] = ''
                image_data['removed_words'] = ''
                data.append(image_data)

    # Now we need to calculate average confidence levels
    average_confidence = total_confidence / images_with_text if images_with_text > 0 else 0
    average_sharpened_confidence = total_sharpened_confidence / total_images if total_images > 0 else 0

    # Return collected data and stats
    return (data, total_images, images_with_text, images_without_text, failed_extractions, low_confidence_count, spell_errors_count, total_extracted_elements, sharpened_low_confidence_count, sharpened_spell_errors_count, total_sharpened_extracted_elements, average_confidence, average_sharpened_confidence, english_texts, mandarin_texts, original_texts)

# Function to save images and text to Excel
# Function to save images and text to Excel
def save_to_excel(directory, data, excel_filename):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    # Set column widths
    ws.column_dimensions['A'].width = 30  # Adjust the width of the column for images
    ws.column_dimensions['B'].width = 20  # Image name column width
    ws.column_dimensions['C'].width = 50  # Extracted text column width
    ws.column_dimensions['D'].width = 15  # Average Confidence column
    ws.column_dimensions['E'].width = 30  # Verification Status column
    ws.column_dimensions['F'].width = 50  # Misspelled Words column width
    ws.column_dimensions['G'].width = 50  # Sharpened Extracted Text
    ws.column_dimensions['H'].width = 15  # Sharpened Avg Confidence
    ws.column_dimensions['I'].width = 30  # Sharpened Verification Status
    ws.column_dimensions['J'].width = 50  # Sharpened Misspelled Words
    ws.column_dimensions['K'].width = 50  # Sharpened Text Without Misspelled Words
    ws.column_dimensions['L'].width = 50  # Added Words
    ws.column_dimensions['M'].width = 50  # Removed Words
    ws.column_dimensions['N'].width = 50  # Mandarin Text from Sharpened Images

    # Write headers
    ws.append(["Image", "Image Name", "Extracted Text", "Avg Confidence", "Verification Status", "Misspelled Words",
               "Sharpened Extracted Text", "Sharpened Avg Confidence", "Sharpened Verification Status",
               "Sharpened Misspelled Words",
               "Sharpened Text Without Misspelled Words", "Added Words", "Removed Words", "Mandarin Text (Sharpened)"])

    for i, image_data in enumerate(data, start=2):
        image_path = image_data['file_path']
        image_name = image_data['filename']

        # Insert image into the Excel sheet
        try:
            img = ExcelImage(image_path)

            # Optionally, adjust image dimensions (optional)
            img.width = 200  # Adjust the width of the image
            img.height = 150  # Adjust the height of the image
            ws.add_image(img, f'A{i}')  # Add image in column A
        except Exception as e:
            print(f"Error adding image {image_name} to Excel: {e}")

        # Write data into the sheet
        ws[f'B{i}'] = image_name
        ws[f'C{i}'] = image_data.get('extracted_text', '')
        ws[f'D{i}'] = round(image_data.get('avg_confidence', 0), 2)
        ws[f'E{i}'] = image_data.get('verification_status', '')
        ws[f'F{i}'] = image_data.get('misspelled_words', '')
        ws[f'G{i}'] = image_data.get('sharpened_extracted_text', '')
        ws[f'H{i}'] = round(image_data.get('sharpened_avg_confidence', 0), 2)
        ws[f'I{i}'] = image_data.get('sharpened_verification_status', '')
        ws[f'J{i}'] = image_data.get('sharpened_misspelled_words', '')
        ws[f'K{i}'] = image_data.get('sharpened_text_without_misspelled', '')
        ws[f'L{i}'] = image_data.get('added_words', '')
        ws[f'M{i}'] = image_data.get('removed_words', '')

        # Add Mandarin text from sharpened image to column N
        mandarin_text_sharpened = extract_mandarin_text(image_data.get('sharpened_extracted_text', ''))
        ws[f'N{i}'] = mandarin_text_sharpened

        # Set row height to match the image height
        row_height = 150 * 0.75  # Approximate conversion from pixels to Excel row height
        ws.row_dimensions[i].height = row_height

    # Save the Excel file in the images directory
    excel_path = os.path.join(directory, excel_filename)
    wb.save(excel_path)
    print(f"Data has been successfully exported to {excel_path}")


# Function to save text files
def save_text_files(directory, english_texts, mandarin_texts, original_texts):
    english_text_file = os.path.join(directory, "english_sharpened_texts.txt")
    mandarin_text_file = os.path.join(directory, "mandarin_texts.txt")
    original_text_file = os.path.join(directory, "original_extracted_texts.txt")

    # Save English texts
    with open(english_text_file, 'w', encoding='utf-8') as f:
        for text in english_texts:
            f.write(text + '\n')
    print(f"English texts saved to {english_text_file}")

    # Save Mandarin texts
    with open(mandarin_text_file, 'w', encoding='utf-8') as f:
        for text in mandarin_texts:
            f.write(text + '\n')
    print(f"Mandarin texts saved to {mandarin_text_file}")

    # Save original extracted texts
    with open(original_text_file, 'w', encoding='utf-8') as f:
        for text in original_texts:
            f.write(text + '\n')
    print(f"Original extracted texts saved to {original_text_file}")

# Function to generate stats report
def generate_stats_report(directory, total_images, images_with_text, images_without_text, failed_extractions, time_taken, low_confidence_count, spell_errors_count, total_extracted_elements, sharpened_low_confidence_count, sharpened_spell_errors_count, total_sharpened_extracted_elements, average_confidence, average_sharpened_confidence):
    report_file_path = os.path.join(directory, "stats_report.txt")

    with open(report_file_path, "w") as report_file:
        report_file.write("EasyOCR Stats Report\n")
        report_file.write("====================\n\n")
        report_file.write(f"Time Taken: {time_taken:.2f} seconds\n")
        report_file.write(f"Total Images: {total_images}\n")
        report_file.write(f"Images with Text Extracted: {images_with_text} ({images_with_text / total_images * 100:.2f}%)\n")
        report_file.write(f"Images without Text: {images_without_text} ({images_without_text / total_images * 100:.2f}%)\n")
        report_file.write(f"Failed Extractions: {failed_extractions} ({failed_extractions / total_images * 100:.2f}%)\n")
        report_file.write(f"Average Confidence Level (Original Images): {average_confidence:.2f}\n")
        report_file.write(f"Average Confidence Level (Sharpened Images): {average_sharpened_confidence:.2f}\n")
        report_file.write(f"Low Confidence Text Elements (Original): {low_confidence_count} ({low_confidence_count} / {total_extracted_elements} = {low_confidence_count / total_extracted_elements * 100:.2f}%)\n")
        report_file.write(f"Low Confidence Text Elements (Sharpened): {sharpened_low_confidence_count} ({sharpened_low_confidence_count} / {total_sharpened_extracted_elements} = {sharpened_low_confidence_count / total_sharpened_extracted_elements * 100:.2f}%)\n")
        report_file.write(f"Spell Errors (English) (Original): {spell_errors_count} ({spell_errors_count} / {total_extracted_elements} = {spell_errors_count / total_extracted_elements * 100:.2f}%)\n")
        report_file.write(f"Spell Errors (English) (Sharpened): {sharpened_spell_errors_count} ({sharpened_spell_errors_count} / {total_sharpened_extracted_elements} = {sharpened_spell_errors_count / total_sharpened_extracted_elements * 100:.2f}%)\n")

    print(f"Stats report saved to {report_file_path}")

# Main function to handle user input and export to Excel
def main():
    # Start the timer
    start_time = time.time()

    # Ask the user for the directory of images
    directory = input("Enter the directory of your images: ").strip()

    if not os.path.isdir(directory):
        print("Invalid directory. Please try again.")
        return

    # Ask the user for the excel file name
    excel_filename = input("Enter the name of your Excel file (without extension): ").strip()
    if not excel_filename:
        excel_filename = "OCR_Results.xlsx"
    else:
        excel_filename += ".xlsx"

    # Extract text from images
    result = extract_text_from_images(directory)
    data = result[0]
    total_images = result[1]
    images_with_text = result[2]
    images_without_text = result[3]
    failed_extractions = result[4]
    low_confidence_count = result[5]
    spell_errors_count = result[6]
    total_extracted_elements = result[7]
    sharpened_low_confidence_count = result[8]
    sharpened_spell_errors_count = result[9]
    total_sharpened_extracted_elements = result[10]
    average_confidence = result[11]
    average_sharpened_confidence = result[12]
    english_texts = result[13]
    mandarin_texts = result[14]
    original_texts = result[15]

    # Save data to Excel
    save_to_excel(directory, data, excel_filename)

    # Save text files
    save_text_files(directory, english_texts, mandarin_texts, original_texts)

    # End the timer
    end_time = time.time()
    time_taken = end_time - start_time

    # Generate stats report
    generate_stats_report(directory, total_images, images_with_text, images_without_text, failed_extractions, time_taken, low_confidence_count, spell_errors_count, total_extracted_elements, sharpened_low_confidence_count, sharpened_spell_errors_count, total_sharpened_extracted_elements, average_confidence, average_sharpened_confidence)

if __name__ == "__main__":
    main()
