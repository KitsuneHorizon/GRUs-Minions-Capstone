"""
PyTorch seemed to run very slow so this might be an impediment. It seemed to get mixed results
whether it extracted more text and even less text. I think it could be very useful if we
pick the optimal OCR method. I could see this library going in either direction (super useful or
not as useful as other pre-processing techniques)
"""

import warnings
warnings.filterwarnings('ignore')

import os
import time
import torch
import torchvision.transforms as transforms
import easyocr
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm  # Import tqdm for progress bar
import re

# Preprocessing function to clean and split text into words
def preprocess_words(text):
    text = re.sub(r'[^\w\s]', '', text).lower()
    words = text.split()
    return words

# Function to perform OCR on an image using two readers
def perform_ocr(reader_simplified, reader_traditional, file_path):
    results = reader_simplified.readtext(file_path, detail=1)

    # Fallback to Traditional Chinese reader if no text is found
    if not results:
        results = reader_traditional.readtext(file_path, detail=1)

    if results:
        extracted_text = ' '.join([res[1] for res in results])
        confidences = [res[2] for res in results]
        avg_confidence = sum(confidences) / len(confidences)
        return extracted_text, avg_confidence, confidences
    else:
        return "No text found", 0, []

# PyTorch-based function to sharpen an image
def sharpen_image_pytorch(image):
    # Convert PIL image to tensor
    transform = transforms.ToTensor()
    image_tensor = transform(image).unsqueeze(0)  # Add batch dimension

    # Define sharpening kernel
    sharpening_kernel = torch.tensor([[
        [-1, -1, -1],
        [-1, 9, -1],
        [-1, -1, -1]
    ]], dtype=torch.float32)

    # Expand kernel dimensions to match image channels
    kernel = sharpening_kernel.expand(image_tensor.size(1), 1, 3, 3)

    # Apply sharpening using convolution
    sharpened_tensor = torch.nn.functional.conv2d(image_tensor, kernel, padding=1, groups=image_tensor.size(1))

    # Clip the values to the range [0, 1]
    sharpened_tensor = torch.clamp(sharpened_tensor, 0, 1)

    # Convert back to PIL image
    transform_to_pil = transforms.ToPILImage()
    return transform_to_pil(sharpened_tensor.squeeze(0))  # Remove batch dimension

# Function to process images and extract text with word count
def extract_text_from_images(directory):
    reader_simplified = easyocr.Reader(['en', 'ch_sim'], verbose=False)
    reader_traditional = easyocr.Reader(['en', 'ch_tra'], verbose=False)

    data = []
    supported_extensions = ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')
    sharpened_image_paths = []  # Track created sharpened image paths

    for filename in tqdm(os.listdir(directory), desc="Processing Images", unit="image"):
        if filename.lower().endswith(supported_extensions):
            file_path = os.path.join(directory, filename)

            # Extract text from original image
            original_text, avg_confidence, _ = perform_ocr(reader_simplified, reader_traditional, file_path)
            original_word_count = len(preprocess_words(original_text))

            # Apply sharpening filter using PyTorch
            try:
                img = Image.open(file_path)
                sharpened_img = sharpen_image_pytorch(img)

                # Save the sharpened image temporarily
                sharpened_image_path = os.path.join(directory, f"sharpened_{filename}")
                sharpened_img.save(sharpened_image_path)
                sharpened_image_paths.append(sharpened_image_path)  # Keep track of the path

                # Extract text from the sharpened image
                sharpened_text, _, _ = perform_ocr(reader_simplified, reader_traditional, sharpened_image_path)
                sharpened_word_count = len(preprocess_words(sharpened_text))
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                sharpened_image_path = None
                sharpened_text = "Error during sharpening"
                sharpened_word_count = 0

            # Append to data
            data.append([
                file_path, original_text, original_word_count, sharpened_image_path,
                sharpened_text, sharpened_word_count
            ])

    return data, sharpened_image_paths

# Function to save images and text to Excel
def save_to_excel(directory, data, excel_filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20

    ws.append([
        "Original Image", "Text from Original Image", "Original Image Word Count",
        "Sharpened Image", "Text from Sharpened Image", "Sharpened Image Word Count"
    ])

    for i, (original_image_path, original_text, original_word_count,
            sharpened_image_path, sharpened_text, sharpened_word_count) in enumerate(data, start=2):

        try:
            orig_img = ExcelImage(original_image_path)
            orig_img.width, orig_img.height = 150, 100
            ws.add_image(orig_img, f'A{i}')
        except Exception as e:
            print(f"Error adding original image to Excel: {e}")

        ws[f'B{i}'] = original_text
        ws[f'C{i}'] = original_word_count

        if sharpened_image_path:
            try:
                sharp_img = ExcelImage(sharpened_image_path)
                sharp_img.width, sharp_img.height = 150, 100
                ws.add_image(sharp_img, f'D{i}')
            except Exception as e:
                print(f"Error adding sharpened image to Excel: {e}")

        ws[f'E{i}'] = sharpened_text
        ws[f'F{i}'] = sharpened_word_count

        row_height = 100 * 0.75
        ws.row_dimensions[i].height = row_height

    excel_path = os.path.join(directory, excel_filename)
    wb.save(excel_path)
    print(f"Data has been successfully exported to {excel_path}")

# Main function to handle user input and export to Excel
def main():
    start_time = time.time()
    directory = input("Enter the directory of your images: ").strip()
    if not os.path.isdir(directory):
        print("Invalid directory. Please try again.")
        return

    excel_filename = input("Enter the name of your Excel file (without extension): ").strip()
    if not excel_filename:
        excel_filename = "OCR_Results.xlsx"
    else:
        excel_filename += ".xlsx"

    # Extract text from images
    data, sharpened_image_paths = extract_text_from_images(directory)

    # Save data to Excel
    save_to_excel(directory, data, excel_filename)

    # Clean up: Delete the temporary sharpened images
    for path in sharpened_image_paths:
        try:
            os.remove(path)
            #print(f"Deleted temporary file: {path}")
        except Exception as e:
            print(f"Error deleting file {path}: {e}")

    end_time = time.time()
    print(f"Time taken: {end_time - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
