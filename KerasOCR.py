# pip install keras==2.9.0 tensorflow==2.12.0


import warnings
warnings.filterwarnings('ignore')

import keras_ocr
import pandas as pd
import os
import time
from spellchecker import SpellChecker
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyXLImage


def extract_text_from_images(directory):
    keras_ocr_pipeline = keras_ocr.pipeline.Pipeline()
    spell_checker = SpellChecker()  # Initialize the spell checker
    images = []
    for filename in os.listdir(directory):
        if filename.endswith(('.png', '.jpg', '.jpeg')):
            images.append(os.path.join(directory, filename))

    data = []
    failed_extractions = 0
    low_confidence_count = 0
    total_extracted_elements = 0
    total_words = 0  # Track total words
    total_misspelled = 0  # Track total misspelled words

    for image_path in images:
        try:
            prediction = keras_ocr_pipeline.recognize([image_path])
            full_text = " ".join([text for text, _ in prediction[0]])
            data.append((image_path, full_text))
            total_extracted_elements += len(prediction[0])

            # Count words and misspelled words
            words = full_text.split()
            total_words += len(words)
            misspelled_words = [word for word in words if word and word not in spell_checker]
            total_misspelled += len(misspelled_words)

            data[-1] += (misspelled_words,)  # Add misspelled words to the data

            if len(full_text) < 3:  # Example condition for low confidence
                low_confidence_count += 1

        except Exception as e:
            print(f"Error processing {image_path}: {e}")
            failed_extractions += 1

    return data, len(images), failed_extractions, low_confidence_count, total_extracted_elements, total_words, total_misspelled


def save_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    ws.append(['Image', 'Extracted Text', 'Misspelled Words Count', 'Misspelled Words'])  # Add headers

    for image_path, extracted_text, misspelled_words in data:
        img = OpenPyXLImage(image_path)
        img.width = 100  # Set the width of the image
        img.height = 100  # Set the height of the image

        img_row = ws.max_row + 1
        img_col = 1

        ws.add_image(img, ws.cell(row=img_row, column=img_col).coordinate)

        ws.row_dimensions[img_row].height = img.height * 0.75  # Adjust for padding
        ws.column_dimensions[
            ws.cell(row=img_row, column=img_col).column_letter].width = img.width * 0.12  # Adjust for padding

        ws.cell(row=img_row, column=2, value=extracted_text)

        # Count and write the number of misspelled words
        misspelled_count = len(misspelled_words)
        ws.cell(row=img_row, column=3, value=misspelled_count)

        # Write the misspelled words as a comma-separated string
        ws.cell(row=img_row, column=4, value=', '.join(misspelled_words))

    wb.save(output_file)


def generate_stats_report(directory, total_images, images_with_text, images_without_text,
                          failed_extractions, time_taken, total_words, total_misspelled):
    report_file_path = os.path.join(directory, 'stats_report.txt')
    with open(report_file_path, 'w') as report_file:
        report_file.write(f"Total Images Processed: {total_images}\n")
        report_file.write(f"Images with Text: {images_with_text}\n")
        report_file.write(f"Images without Text: {images_without_text}\n")
        report_file.write(f"Failed Extractions: {failed_extractions}\n")
        report_file.write(f"Total Words: {total_words}\n")
        report_file.write(f"Total Misspelled Words: {total_misspelled}\n")
        report_file.write(f"Time Taken: {time_taken:.2f} seconds\n")


def main():
    directory = input("Enter the directory of your images: ")
    output_file_name = input("Enter the name of your Excel file (without extension): ") + '.xlsx'
    output_file_path = os.path.join(directory, output_file_name)

    start_time = time.time()

    data, total_images, failed_extractions, low_confidence_count, total_extracted_elements, total_words, total_misspelled = extract_text_from_images(directory)

    save_to_excel(data, output_file_path)

    time_taken = time.time() - start_time
    images_with_text = len(data)
    images_without_text = total_images - images_with_text

    generate_stats_report(directory, total_images, images_with_text, images_without_text,
                          failed_extractions, time_taken, total_words, total_misspelled)

    print(f"Data has been successfully exported to {output_file_path}")
    print(
        f"Total Images: {total_images}, Failed Extractions: {failed_extractions}, Time Taken: {time_taken:.2f} seconds")


if __name__ == "__main__":
    main()
