import os
from PIL import Image
import pytesseract
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# Update the path to your Tesseract-OCR executable
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Update this path as needed

def resize_image(image_path, max_width, max_height):
    """Resize image to fit within max_width x max_height while preserving aspect ratio."""
    try:
        with Image.open(image_path) as img:
            img.thumbnail((max_width, max_height))
            return img
    except Exception as e:
        print(f"Error resizing image {os.path.basename(image_path)}: {e}")
        return None

def extract_text_from_image(image_path):
    """Extract text from an image using pytesseract OCR."""
    try:
        with Image.open(image_path) as img:
            text = pytesseract.image_to_string(img)
            return text.strip()
    except Exception as e:
        print(f"Cannot extract text from image {os.path.basename(image_path)}. Error: {e}")
        return ""

def get_output_file_name(image_directory, output_file_name):
    """Prompt the user to either overwrite the existing file or choose a new name."""
    while True:
        output_excel_file = os.path.join(image_directory, f'{output_file_name}.xlsx')
        if os.path.exists(output_excel_file):
            print(f"File '{output_file_name}.xlsx' already exists.")
            choice = input("Do you want to overwrite it (o) or choose a new name (n)? ").strip().lower()
            if choice == 'o':
                return output_excel_file
            elif choice == 'n':
                output_file_name = input("Enter a new name for the Excel file (without extension): ").strip()
            else:
                print("Invalid choice. Please enter 'o' to overwrite or 'n' to choose a new name.")
        else:
            return output_excel_file

def images_to_excel(image_directory, output_file_name):
    # Get the output file path with user choice to overwrite or rename
    output_excel_file = get_output_file_name(image_directory, output_file_name)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Set column widths
    ws.column_dimensions['A'].width = 30  # Images column
    ws.column_dimensions['B'].width = 20  # File Names column
    ws.column_dimensions['C'].width = 25  # Image Dimensions column
    ws.column_dimensions['D'].width = 15  # Formats column
    ws.column_dimensions['E'].width = 50  # Extracted Text column
    ws.column_dimensions['F'].width = 15  # Status column

    # Add headers
    ws['A1'] = 'Images'
    ws['B1'] = 'File Names'
    ws['C1'] = 'Image Dimensions'
    ws['D1'] = 'Formats'
    ws['E1'] = 'Extracted Text'
    ws['F1'] = 'Status'

    # Initial row for data
    row = 2
    temp_files = []  # List to keep track of temporary files

    # Get a list of image files in the directory
    try:
        image_files = [f for f in os.listdir(image_directory) if
                       f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
        if not image_files:
            print("No images found in the directory.")
            return
    except PermissionError:
        print(f"Permission denied to access directory: {image_directory}")
        return
    except FileNotFoundError:
        print(f"Directory not found: {image_directory}")
        return

    for image_file in image_files:
        # Construct the full file path
        image_path = os.path.join(image_directory, image_file)

        # Define the temporary image file path
        temp_img_path = os.path.join(image_directory, f'temp_{image_file}')
        temp_files.append(temp_img_path)  # Add to list of temporary files

        try:
            # Resize the image
            img = resize_image(image_path, 180, 100)  # Adjust max_width and max_height as needed

            if img is None:
                print(f"Image is corrupted or cannot be opened: {image_file}")
                # Log data for corrupted images
                ws[f'B{row}'] = image_file
                ws[f'C{row}'] = 'NA'
                ws[f'D{row}'] = 'NA'
                ws[f'E{row}'] = 'NA'
                ws[f'F{row}'] = 'Corrupted'
                row += 1
                continue

            # Save the image temporarily for openpyxl to use
            img.save(temp_img_path)
            print(f"Processing: {image_file}")  # Changed print statement to show only the image name
        except (IOError, SyntaxError, AttributeError) as e:
            print(f"Cannot open or save image file: {image_file}. Error: {e}")
            # Log data for corrupted images
            ws[f'B{row}'] = image_file
            ws[f'C{row}'] = 'NA'
            ws[f'D{row}'] = 'NA'
            ws[f'E{row}'] = 'NA'
            ws[f'F{row}'] = 'Corrupted'
            row += 1
            continue

        # Extract text from the image
        text = extract_text_from_image(image_path)

        # Extract image dimensions and format
        img_width, img_height = img.size
        img_format = img.format

        # Add image to the Excel sheet
        try:
            excel_image = XLImage(temp_img_path)
            excel_image.width = img.width
            excel_image.height = img.height
            ws.add_image(excel_image, f'A{row}')

            # Add file name, dimensions, format, and extracted text to the Excel sheet
            ws[f'B{row}'] = image_file
            ws[f'C{row}'] = f'{img_width}x{img_height}'
            ws[f'D{row}'] = img_format
            ws[f'E{row}'] = text
            ws[f'F{row}'] = 'OK'

            # Adjust row height based on image height
            ws.row_dimensions[row].height = img.height * 0.75  # Adjust factor as needed for padding
        except Exception as e:
            print(f"Cannot add image to Excel sheet: {image_file}. Error: {e}")
            ws[f'B{row}'] = image_file
            ws[f'C{row}'] = f'{img_width}x{img_height}'
            ws[f'D{row}'] = img_format
            ws[f'E{row}'] = text
            ws[f'F{row}'] = 'Error Adding Image'

        # Move to the next row
        row += 1

    # Save the workbook
    try:
        wb.save(output_excel_file)
        print(f"Excel file created successfully: {os.path.basename(output_excel_file)}")
    except IOError as e:
        print(f"Cannot save Excel file: {os.path.basename(output_excel_file)}. Error: {e}")

    # Remove temporary image files after saving the workbook
    for temp_file in temp_files:
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except IOError as e:
                print(f"Cannot remove temporary image file: {os.path.basename(temp_file)}. Error: {e}")

    # Print completion message
    print("Tool Complete -- See Excel File")

if __name__ == "__main__":
    # Take user input for the image directory
    image_directory = input("Enter the path to the image directory: ").strip()

    # Validate the directory input
    if not os.path.isdir(image_directory):
        print("The specified directory does not exist or is not accessible.")
    else:
        # Prompt for the desired Excel file name
        output_file_name = input("Enter the desired name for the Excel file (without extension): ").strip()
        images_to_excel(image_directory, output_file_name)
