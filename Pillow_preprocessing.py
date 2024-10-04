
"""
PILLOW Analysis:
PILLOW has been really good about extracting more words from the images after sharpening.
I think this one has potential to be used as a preprocessing methodology with the optimal OCR pairing
"""
import warnings
warnings.filterwarnings('ignore')

import os
import time
import easyocr
from PIL import Image, ImageFilter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm  # Import tqdm for progress bar
import re

# Preprocessing function to clean and split text into words
def preprocess_words(text):
    # Remove punctuation using regex and convert to lowercase
    text = re.sub(r'[^\w\s]', '', text).lower()
    # Split text into a list of words
    words = text.split()
    return words

# Function to perform OCR on an image using two readers
def perform_ocr(reader_simplified, reader_traditional, file_path):
    results = reader_simplified.readtext(file_path, detail=1)

    # Fallback to Traditional Chinese reader if no text is found
    if not results:
        results = reader_traditional.readtext(file_path, detail=1)

    if results:
        extracted_text = ' '.join([res[1] for res in results])
        confidences = [res[2] for res in results]
        avg_confidence = sum(confidences) / len(confidences)
        return extracted_text, avg_confidence, confidences
    else:
        return "No text found", 0, []

# Function to process images and extract text with word count
def extract_text_from_images(directory):
    # Initialize the OCR readers
    reader_simplified = easyocr.Reader(['en', 'ch_sim'], verbose=False)
    reader_traditional = easyocr.Reader(['en', 'ch_tra'], verbose=False)

    # Initialize list to store data for the Excel file
    data = []

    supported_extensions = ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')

    # Loop through the directory with tqdm progress bar
    for filename in tqdm(os.listdir(directory), desc="Processing Images", unit="image"):
        if filename.lower().endswith(supported_extensions):
            file_path = os.path.join(directory, filename)

            # Extract text from original image
            original_text, avg_confidence, _ = perform_ocr(reader_simplified, reader_traditional, file_path)
            original_word_count = len(preprocess_words(original_text))  # Calculate word count

            # Apply sharpening filter using Pillow
            try:
                img = Image.open(file_path)
                sharpened_img = img.filter(ImageFilter.SHARPEN)

                # Save the sharpened image temporarily
                sharpened_image_path = os.path.join(directory, f"sharpened_{filename}")
                sharpened_img.save(sharpened_image_path)

                # Extract text from the sharpened image
                sharpened_text, _, _ = perform_ocr(reader_simplified, reader_traditional, sharpened_image_path)
                sharpened_word_count = len(preprocess_words(sharpened_text))  # Calculate word count
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                sharpened_image_path = None
                sharpened_text = "Error during sharpening"
                sharpened_word_count = 0

            # Append to data
            data.append([
                file_path, original_text, original_word_count, sharpened_image_path,
                sharpened_text, sharpened_word_count
            ])

    return data

# Function to save images and text to Excel
def save_to_excel(directory, data, excel_filename):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    # Set column widths
    ws.column_dimensions['A'].width = 30  # Original Image
    ws.column_dimensions['B'].width = 50  # Text from Original Image
    ws.column_dimensions['C'].width = 20  # Original Image Word Count
    ws.column_dimensions['D'].width = 30  # Sharpened Image
    ws.column_dimensions['E'].width = 50  # Text from Sharpened Image
    ws.column_dimensions['F'].width = 20  # Sharpened Image Word Count

    # Write headers
    ws.append([
        "Original Image", "Text from Original Image", "Original Image Word Count",
        "Sharpened Image", "Text from Sharpened Image", "Sharpened Image Word Count"
    ])

    for i, (original_image_path, original_text, original_word_count,
            sharpened_image_path, sharpened_text, sharpened_word_count) in enumerate(data, start=2):

        # Insert original image
        try:
            orig_img = ExcelImage(original_image_path)
            orig_img.width, orig_img.height = 150, 100  # Adjust dimensions
            ws.add_image(orig_img, f'A{i}')
        except Exception as e:
            print(f"Error adding original image to Excel: {e}")

        # Write text and word count from the original image
        ws[f'B{i}'] = original_text
        ws[f'C{i}'] = original_word_count

        # Insert sharpened image if available
        if sharpened_image_path:
            try:
                sharp_img = ExcelImage(sharpened_image_path)
                sharp_img.width, sharp_img.height = 150, 100  # Adjust dimensions
                ws.add_image(sharp_img, f'D{i}')
            except Exception as e:
                print(f"Error adding sharpened image to Excel: {e}")

        # Write text and word count from the sharpened image
        ws[f'E{i}'] = sharpened_text
        ws[f'F{i}'] = sharpened_word_count

        # Set row height to match the image height
        row_height = 100 * 0.75  # Approximate conversion from pixels to Excel row height
        ws.row_dimensions[i].height = row_height

    # Save the Excel file in the images directory
    excel_path = os.path.join(directory, excel_filename)
    wb.save(excel_path)
    print(f"Data has been successfully exported to {excel_path}")

# Main function to handle user input and export to Excel
def main():
    start_time = time.time()

    directory = input("Enter the directory of your images: ").strip()
    if not os.path.isdir(directory):
        print("Invalid directory. Please try again.")
        return

    # Ask for the Excel file name
    excel_filename = input("Enter the name of your Excel file (without extension): ").strip()
    if not excel_filename:
        excel_filename = "OCR_Results.xlsx"
    else:
        excel_filename += ".xlsx"

    # Extract text from images
    data = extract_text_from_images(directory)

    # Save data to Excel
    save_to_excel(directory, data, excel_filename)

    # End the timer
    end_time = time.time()
    print(f"Time taken: {end_time - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
